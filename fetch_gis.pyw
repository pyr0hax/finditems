from arcgis.gis import GIS
from decimal import Decimal
from openpyxl import Workbook
import concurrent.futures
from datetime import datetime
from configparser import ConfigParser
from os import path
from sys import exit
import tkinter as tk
from tkinter import messagebox

nowdate = datetime.now().strftime("%Y/%m/%d:%H:%M:%S")
configFile = 'config.ini'

if path.exists(configFile):
    config = ConfigParser()
    config.read(configFile)
else:
    with open(configFile, 'a') as f:
        f.write("""[credentials]
host = https://portal.example.com/webadaptor
user = username
passwd = password""")
        root = tk.Tk()
        root.overrideredirect(1)
        root.withdraw()
        messagebox.showinfo("Information", "config.ini file has been created. Please modify to correct settings pointing to your portal and run program again.")
        root.destroy()
        exit()

host = config['credentials']['host']
user = config['credentials']['user']
passwd = config['credentials']['passwd']

gis = GIS(host, user, passwd)

content = gis.content.search(query="", max_items=10000)

workbook = Workbook()
sheet = workbook.active
sheet.title = "Portal Content"

sheet["A1"] = "Item ID"
sheet["B1"] = "Item Name"
sheet["C1"] = "Item Owner"
sheet["D1"] = "Item Size (MB)"

counter = 0

def process_item(item, row):
    global counter
    if item.owner not in ['esri_apps', 'esri_nav', 'esri']:
        sheet.cell(row=row, column=1, value=item.id)
        sheet.cell(row=row, column=2, value=item.title)
        sheet.cell(row=row, column=3, value=item.owner)
        sheet.cell(row=row, column=4, value=Decimal(item.size)/ 1000000)
        counter += 1
        with open('logger.txt', 'a') as f:
            f.write(nowdate + ': ' + item.title + ' ' + 'Number: '  + str(counter) + '\n')

num_threads = 8

with concurrent.futures.ThreadPoolExecutor(max_workers=num_threads) as executor:
    futures = []
    row = 2
    for item in content:
        if item.owner not in ['esri_apps', 'esri_nav', 'esri']:
            futures.append(executor.submit(process_item, item, row))
            row += 1
        else:
            with open('logger.txt', 'a') as f:
                f.write(nowdate + ': ' + item.title + ' ' + 'Skipped' + '\n')
    concurrent.futures.wait(futures)

workbook.save("output.xlsx")
